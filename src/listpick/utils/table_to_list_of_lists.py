#!/bin/python
# -*- coding: utf-8 -*-
"""
table_to_lists.py

Author: GrimAndGreedy
License: MIT
"""

import sys
import csv
import json
from io import StringIO
import argparse
from typing import Tuple, Iterable, Optional
import dill as pickle
import os

def read_file_content(file_path: str) -> str:
    """ Read lines from file. """
    with open(file_path, 'r') as file:
        return file.read()

def strip_whitespace(item: Iterable) -> Iterable:
    """ Strip whitespace from string or from list of strings. """
    if isinstance(item, list):
        return [strip_whitespace(sub_item) for sub_item in item]
    elif isinstance(item, str):
        return item.strip()
    else:
        return item



def table_to_list(input_arg: str, delimiter:str='\t', file_type:Optional[str]=None) -> Tuple[list[list[str]], list[str]]:
    """ 
    Convert data string to list. The input_arg
    Currently accepts: csv, tsv, json, xlsx, ods
    """
    table_data = []

    def parse_csv_like(data:str, delimiter:str) -> list[list[str]]:
        """ Convert value-separated data (e.g., CSV or TSV) to list of lists. """

        try:
            reader = csv.reader(StringIO(data), delimiter=delimiter)
            return [row for row in reader]
        except Exception as e:
            print(f"Error reading CSV-like input: {e}")
            return []

    def csv_string_to_list(csv_string:str) -> list[list[str]]:
        """ Convert csv string to list of lists using csv.reader. """
        f = StringIO(csv_string)
        reader = csv.reader(f, skipinitialspace=True)
        return [row for row in reader]

    if file_type == 'csv' or delimiter in [',']:
        try:
            if input_arg == '--stdin':
                input_data = sys.stdin.read()
            elif input_arg == '--stdin2':
                input_count = int(sys.stdin.readline())
                input_data = "\n".join([sys.stdin.readline().strip() for i in range(input_count)])
            else:
                input_data = read_file_content(input_arg)

            table_data = csv_string_to_list(input_data)
            table_data = strip_whitespace(table_data)
            # table_data = parse_csv_like(input_data, ",")
            return table_data, []
        except Exception as e:
            print(f"Error reading CSV/TSV input: {e}")
            return [], []

    elif file_type == 'tsv':
        try:
            if input_arg == '--stdin':
                input_data = sys.stdin.read()
            elif input_arg == '--stdin2':
                input_count = int(sys.stdin.readline())
                input_data = "\n".join([sys.stdin.readline().strip() for i in range(input_count)])
            else:
                input_data = read_file_content(input_arg)
            
            # Adjust delimiter for TSV or CSV
            if file_type == 'tsv' or delimiter == '\t':
                delimiter = '\t'
            else:
                delimiter = ','
            
            table_data = parse_csv_like(input_data, delimiter)
            table_data = strip_whitespace(table_data)
            return table_data, []
        except Exception as e:
            print(f"Error reading CSV/TSV input: {e}")
            return [], []

    elif file_type == 'json':
        try:
            if input_arg == '--stdin':
                input_data = sys.stdin.read()
            elif input_arg == '--stdin2':
                input_count = int(sys.stdin.readline())
                input_data = "\n".join([sys.stdin.readline() for i in range(input_count)])
            else:
                input_data = read_file_content(input_arg)
            
            table_data = json.loads(input_data)
            return table_data, []
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON input: {e}")
            return [], []
        except FileNotFoundError as e:
            print(f"File not found: {e}")
            return [], []

    elif file_type == 'xlsx':
        from openpyxl import load_workbook
        try:
            if input_arg == '--stdin':
                input_data = sys.stdin.read()
                with open('temp.xlsx', 'wb') as f:
                    f.write(input_data.encode())
            elif input_arg == '--stdin2':
                input_count = int(sys.stdin.readline())
                input_data = "\n".join([sys.stdin.readline() for i in range(input_count)])
                with open('temp.xlsx', 'wb') as f:
                    f.write(input_data.encode())
            else:
                input_data = read_file_content(input_arg)
                with open('temp.xlsx', 'wb') as f:
                    f.write(input_data.encode())
            
            wb = load_workbook(filename='temp.xlsx')
            sheet = wb.active
            for row in sheet.iter_rows(values_only=True):
                table_data.append(list(row))
            return table_data, []
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return [], []

    elif file_type == 'ods':
        try:
            import pandas as pd
            df = pd.read_excel(input_arg, engine='odf')
            table_data = df.values.tolist()
            return table_data, []
        except Exception as e:
            print(f"Error loading ODS file: {e}")
            return [], []
    elif file_type == 'pkl':
        with open(os.path.expandvars(os.path.expanduser(input_arg)), 'rb') as f:
            loaded_data = pickle.load(f)
        items = loaded_data["items"] if "items" in loaded_data else []
        header = loaded_data["header"] if "header" in loaded_data else []
        return items, header

    if input_arg == '--stdin':
        input_data = sys.stdin.read()
    elif input_arg == '--stdin2':
        input_count = int(sys.stdin.readline())
        input_data = "\n".join([sys.stdin.readline() for i in range(input_count)])
    else:
        input_data = read_file_content(input_arg)
    
    table_data = parse_csv_like(input_data, delimiter)

    return table_data, []

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Convert table to list of lists.')
    parser.add_argument('-i', dest='file', help='File containing the table to be converted')
    parser.add_argument('--stdin', action='store_true', help='Table passed on stdin')
    parser.add_argument('--stdin2', action='store_true', help='Table passed on stdin')
    parser.add_argument('-d', dest='delimiter', default='\t', help='Delimiter for rows in the table (default: tab)')
    parser.add_argument('-t', dest='file_type', choices=['tsv', 'csv', 'json', 'xlsx', 'ods'], help='Type of file (tsv, csv, json, xlsx, ods)')
    
    args = parser.parse_args()
    
    if args.file:
        input_arg = args.file
    elif args.stdin:
        input_arg = '--stdin'
    elif args.stdin2:
        input_arg = '--stdin2'
    else:
        print("Error: Please provide input file or use --stdin option.")
        sys.exit(1)
    
    table_data = table_to_list(input_arg, args.delimiter, args.file_type)
    # print(table_data)

    len(table_data[0])
    for row in table_data:
        if len(row) != len(table_data[0]):
            print(len(row), row)
